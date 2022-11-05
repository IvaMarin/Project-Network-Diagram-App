import sys, math, os
import numpy as np
from os import listdir
from os.path import isfile, join
import openpyxl

from PyQt5 import QtWidgets, QtGui ,QtCore
from PyQt5.QtCore import QRect
from PyQt5.QtGui import QIntValidator
from PyQt5.QtWidgets import QApplication, QAction, QMessageBox, QDialog
#############    Первый запуск    #######################
from first_launch.win_first_launch import Ui_first_launch
from encrypt_module import initial_decrypt_file, aes_encrypt, aes_generate_key
from message_box_creator import message_box_create
from pathlib import Path
############################################################
########################     UI     #########################
from qt_designer_ui.login import Ui_login
from qt_designer_ui.startWindow import Ui_startWin
from qt_designer_ui.winEditTable import Ui_CreatEditTask
from qt_designer_ui.EditTable import Ui_Dialog
from qt_designer_ui.tableNumPeopleInSquad import Ui_winTableNumPeopleInSquad
from qt_designer_ui.setNumSquad import Ui_SetNumSquad
#############################################################


def find_files(catalog: Path):

    res = []
    for root, dirs, files in os.walk(catalog.resolve()):
        res += [Path(os.path.join(root, name)) for name in files
                if Path(name).suffix.lower() in ['.json', '.jpg', '.txt']]
    return res


class winSigReport(QtWidgets.QDialog): # окно изменения личных данных студента в запущенном приложении (доступно только преподавателю)

    def __init__(self, root): # передаем параметр root это родитель т е MainMenu (в этом классе и лежит наше окно winSigReport)
        """Initializer."""
        super().__init__(root) # инициализация

        self.ui = Ui_login() # инициализация ui
        self.ui.setupUi(self) # инициализация ui окна (присвоение конкретных пар-ов)
        self.mainMenu = root  # сохраняем нашего родителя

        self.ui.lineEditNumINGroup.setValidator(QIntValidator())

        rx = QtCore.QRegExp("[a-zA-Zа-яА-Я .,]{100}")
        val = QtGui.QRegExpValidator(rx)
        self.ui.lineEditSurname.setValidator(val)

        sizeWindow = QRect(QApplication.desktop().screenGeometry())         # смотрим размер экраны
        width = int(sizeWindow.width() - (sizeWindow.width()) * 2 / 3)      # выставляем ширину окна
        height = int(sizeWindow.height() - (sizeWindow.height()) * 2 / 3)   # выставляем длину окна
        # присваиваем параметры длины и ширины окну
        self.resize(width, height)

        self.move(int(sizeWindow.width() / 20), int(sizeWindow.height() / 20)) # двигаем окно левее и выше

        self.ui.lineEditSurname.insert(self.mainMenu.surname)       # подгружаем из mainMenu данные если они уже были указаны
        self.ui.lineEditNumINGroup.insert(self.mainMenu.numINGroup) #
        self.ui.lineEditGroup.insert(self.mainMenu.numGroup)        #

        self._connectAction() # ф-ия связи с эл-тами окна

    def _connectAction(self):
        self.ui.btnSignLab.clicked.connect(lambda: self.saveData()) # прописываем действие по кнопке

    def saveData(self): # сохраняем имя фамилию и № группы полученные в этом диалоговом окне
        if self.checkInputData() : # проверка входящих данных
            return
        else:
            self.mainMenu.surname = self.ui.lineEditSurname.text()  # сохраняем в класс WindowMenu фамилию
            self.mainMenu.numINGroup = self.ui.lineEditNumINGroup.text()  # сохраняем в класс WindowMenu группу
            self.mainMenu.numGroup = self.ui.lineEditGroup.text()  # сохраняем в класс WindowMenu группу
            # WindowMenu это класс окна Меню

            #self.mainMenu.creatReport() # перезапись в pdf данных студента

            self.close()

    def checkInputData(self): # проверка входящих данных (есть ли незаполненные строки или
        # существует ли файл с указанным номером варианта )
        fileName = "В" + self.ui.lineEditNumINGroup.text() + ".xlsx"
        pathFileXlsx = os.path.join("resources", "variants", fileName)

        if self.ui.lineEditSurname.text() == "" or\
                self.ui.lineEditNumINGroup.text() == "" or\
                self.ui.lineEditGroup.text() == "": # если существует незаполненная строка, выводим предупреждение и
            # возврахаем True чтобы сработало условие в функции откуда вызывалась данная функция
            warning = QMessageBox()
            warning.setWindowTitle("Предупреждение")
            warning.setText("Заполните все предложенные поля.")
            warning.setDefaultButton(QMessageBox.Ok)
            warning = warning.exec()
            return True
        elif not(os.path.exists(pathFileXlsx)): # если не существует файла с указанным вариантом, выводим предупреждение и
            # возврахаем True чтобы сработало условие в функции откуда вызывалась данная функция
            warning = QMessageBox()
            warning.setWindowTitle("Предупреждение")
            warning.setText("Введите корректный номер варианта.")
            warning.setDefaultButton(QMessageBox.Ok)
            warning = warning.exec()
            return True
        else: # иначе возвращаем False (проверки пройдены)
            return False

class winLogin(QtWidgets.QDialog):# Окно регистрации в приложении

    def __init__(self, root): # передаем параметр root это родитель т е MainMenu (в этом классе и лежит наше окно winSigReport)
        """Initializer."""
        super().__init__(root) # инициализация

        self.ui = Ui_startWin() # инициализация ui
        self.ui.setupUi(self) # инициализация ui окна (присвоение конкретных пар-ов)
        self.mainMenu = root  # сохраняем нашего родителя

        self.ui.lineEditNumINGroup.setValidator(QIntValidator())

        rx = QtCore.QRegExp("[a-zA-Zа-яА-Я .,]{100}")
        val = QtGui.QRegExpValidator(rx)
        self.ui.lineEditSurname.setValidator(val)

        sizeWindow = QRect(QApplication.desktop().screenGeometry())         # смотрим размер экраны
        width = int(sizeWindow.width())      # выставляем ширину окна
        height = int(sizeWindow.height())   # выставляем длину окна
        # присваиваем параметры длины и ширины окну
        self.resize(width, height)

        #self.move(int(sizeWindow.width() / 20), int(sizeWindow.height() / 20)) # двигаем окно левее и выше

        self._connectAction() # ф-ия связи с эл-тами окна

        quit = QAction("Quit", self) # событие выхода
        quit.triggered.connect(self.closeEvent) # если событие выхода срабатывает то вызывается closeEvent

    def closeEvent(self, event):
        if (self.ui.btnSignLab.isChecked() and self.checkInputData()): # если кнопка подписи отчета нажата и проверка входящих данных True
            event.ignore() # оставляем текущее окно открытым
        elif self.ui.btnSignLab.isChecked(): # если closeEvent вызван и при этом нажата кнопка подписи отчета
            event.accept() # то не выводим диалоговое окно подтверждения выхода из проги
        else: # иначе формируем окно подтверждения выхода из проги (т.е QMessageBox)
            close = QMessageBox()
            close.setWindowTitle("Закрыть приложение")
            close.setText("Вы уверены, что хотите закрыть приложение?") #
            close.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel) #
            close = close.exec()
            if close == QMessageBox.Ok: # если нажали да
                event.accept() # подтверждаем ивент
                sys.exit()
            else: # иначе игнорируем
                event.ignore()
        self.ui.btnSignLab.setChecked(False)

    def checkInputData(self):# проверка входящих данных (есть ли незаполненные строки или
        # существует ли файл с указанным номером варианта )
        fileName = "В" + self.ui.lineEditNumINGroup.text() + ".xlsx"
        pathFileXlsx = os.path.join("resources", "variants", fileName)

        if self.ui.lineEditSurname.text() == "" or\
                self.ui.lineEditNumINGroup.text() == "" or\
                self.ui.lineEditGroup.text() == "":# если существует незаполненная строка, выводим предупреждение и
            # возврахаем True чтобы сработало условие в функции откуда вызывалась данная функция
            warning = QMessageBox()
            warning.setWindowTitle("Предупреждение")
            warning.setText("Заполните все предложенные поля.")
            warning.setDefaultButton(QMessageBox.Ok)
            warning = warning.exec()
            return True
        elif not(os.path.exists(pathFileXlsx)):# если не существует файла с указанным вариантом, выводим предупреждение и
            # возврахаем True чтобы сработало условие в функции откуда вызывалась данная функция
            warning = QMessageBox()
            warning.setWindowTitle("Предупреждение")
            warning.setText("Введите корректный номер варианта.")
            warning.setDefaultButton(QMessageBox.Ok)
            warning = warning.exec()
            return True
        else: # иначе возвращаем False (проверки пройдены)
            return False

    def _connectAction(self):
        self.ui.btnSignLab.clicked.connect(lambda: self.saveData()) # прописываем действие по кнопке
        self.ui.btnDeveloperMode.clicked.connect(lambda: self.activateDeveloperMode()) #

    def activateDeveloperMode (self):
        self.mainMenu.activateDeveloperMode()

        self.ui.lineEditSurname.insert(self.mainMenu.surname)        # подгружаем из mainMenu данные если они уже были указаны
        self.ui.lineEditNumINGroup.insert(self.mainMenu.numINGroup)  #
        self.ui.lineEditGroup.insert(self.mainMenu.numGroup)         #


    def saveData(self): # сохраняем имя фамилию и № группы полученные в этом диалоговом окне
        self.mainMenu.surname = self.ui.lineEditSurname.text()  # сохраняем в класс WindowMenu фамилию
        self.mainMenu.numINGroup = self.ui.lineEditNumINGroup.text()# сохраняем в класс WindowMenu группу
        self.mainMenu.numGroup = self.ui.lineEditGroup.text()  # сохраняем в класс WindowMenu группу
        # WindowMenu это класс окна Меню
        self.close()


class winEditTable(QtWidgets.QDialog): # окно выбора файлов с таблицами, для редактирования 
    def __init__(self, root):  # передаем параметр root это родитель т е MainMenu (в этом классе и лежит наше окно winSigReport)
        """Initializer."""
        super().__init__(root)  # инициализация

        self.ui = Ui_CreatEditTask()  # инициализация ui
        self.ui.setupUi(self)  # инициализация ui окна (присвоение конкретных пар-ов)
        self.mainMenu = root  # сохраняем нашего родителя
        self.fileName = ""

        sizeWindow = QRect(QApplication.desktop().screenGeometry())  # смотрим размер экраны
        width = int(sizeWindow.width() - (sizeWindow.width()) * 2 / 3)  # выставляем ширину окна
        height = int(sizeWindow.height() - (sizeWindow.height()) * 2 / 3)  # выставляем длину окна
        # присваиваем параметры длины и ширины окну
        self.resize(width, height)


        self.move(int(sizeWindow.width() / 20), int(sizeWindow.height() / 20))  # двигаем окно левее и выше

        self.ui.lineEdit.setValidator(QIntValidator())
        self.ui.lineEdit.setMaxLength(2)

        pathFileXlsx = os.path.join("resources", "variants")  # находим путь до папки с файлами вариантов
        self.onlyfiles = [f for f in listdir(pathFileXlsx) if isfile(join(pathFileXlsx, f))] # собираем список всех файлов в этой папке
        self.ui.comboBoxVariants.addItems([name for name in self.onlyfiles]) # загружаем список файлов в comboBoxVariants

        self.creatTable = creatTable(self)  # создаем окно с таблицей для редактирования вариантов
        #fileName = self.ui.comboBoxVariants

        self._connectAction()  # ф-ия связи с эл-тами окна

    def _connectAction(self):
        self.ui.btnCreatTable.clicked.connect(lambda: self.creatNewTable())
        self.ui.btnEditTable.clicked.connect(lambda: self.editTable())
        self.ui.btnDeletTable.clicked.connect(lambda: self.deleteTable())

    def creatNewTable(self): # функция создания файлов с вариантами таблиц для лабы
        newTableVar = openpyxl.Workbook() # создание книги (Excel файла)
        self.fileName = "В" + self.ui.lineEdit.text() + ".xlsx" # генерируем имя Excel файла
        pathFileXlsx = os.path.join("resources", "variants", self.fileName)  # находим путь до директории с вариантами
        # проверяем существует ли файл с указанным названием (self.fileName) по пути pathFileXlsx
        if os.path.isfile(pathFileXlsx): # если существует
            warning = QMessageBox()  # выводим предупреждение
            warning.setWindowTitle("Предупреждение")
            warning.setText("Такой файл уже существует.")  #
            warning.setDefaultButton(QMessageBox.Ok)  #
            warning = warning.exec()  #
        else: # иначе сохраняем созданную пустую книгу с названием файла self.fileName по пути в директорию pathFileXlsx
            newTableVar.save(pathFileXlsx) #
            self.ui.comboBoxVariants.addItem(self.fileName) # добавляем название файла в выпадающий список

            self.close() #
            self.creatTable.openFile(pathFileXlsx) #
            self.creatTable.exec_()

    def editTable(self):
        #self.ui.comboBoxVariants.currentText()  выбранный вариант из comboBoxVariants
        self.fileName = os.path.join("resources", "variants", self.ui.comboBoxVariants.currentText())  # находим путь до файла

        self.close()
        self.creatTable.openFile(self.fileName) # открываем указанный файл в окне для редактирования вариантов
        self.creatTable.exec_()

    def deleteTable(self): # удаление файла с вариантом и удаление его названия из выпадающего списка
        self.fileName = os.path.join("resources", "variants",
                                     self.ui.comboBoxVariants.currentText())  # находим путь до файла
        try:
            with open(self.fileName, "r") as file:
            # Распечатать сообщение об успешном завершении
                file.close()
            close = QMessageBox()
            close.setWindowTitle("Удалить вариант")
            close.setText("Вы уверены, что хотите удалить вариант?")  #
            close.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)  #
            close = close.exec()
            if close == QMessageBox.Ok:  # если нажали да
                os.remove(self.fileName)
                self.ui.comboBoxVariants.removeItem(self.ui.comboBoxVariants.currentIndex())
            else:  # иначе игнорируем
                return

        # Вызовите ошибку, если файл был открыт раньше
        except OSError:
            warning = QMessageBox()  # выводим предупреждение
            warning.setWindowTitle("Предупреждение")
            warning.setText("Закройте выбранный файл.")  #
            warning.setDefaultButton(QMessageBox.Ok)  #
            warning = warning.exec()  #

class creatTable(QtWidgets.QDialog): # окно с таблицей для непосредственного ее редактирования 
    def __init__(self,root):  # передаем параметр root это родитель т е MainMenu (в этом классе и лежит наше окно winSigReport)
        """Initializer."""
        super().__init__(root)  # инициализация

        self.ui = Ui_Dialog()  # инициализация ui
        self.ui.setupUi(self)  # инициализация ui окна (присвоение конкретных пар-ов)
        self.winEditTable = root  # сохраняем нашего родителя

        sizeWindow = QRect(QApplication.desktop().screenGeometry())  # смотрим размер экраны
        width = int(sizeWindow.width() - (sizeWindow.width()) / 3)  # выставляем ширину окна
        height = int(sizeWindow.height() - (sizeWindow.height()) / 3)  # выставляем длину окна
        # присваиваем параметры длины и ширины окну
        self.resize(width, height)

        self.move(int(sizeWindow.width() / 20), int(sizeWindow.height() / 20))  # двигаем окно левее и выше

        quit = QAction("Quit", self)  # событие выхода
        quit.triggered.connect(self.closeEvent)  # если событие выхода срабатывает то вызывается closeEvent

        self.listNumPeopleInSquad = []
        self._connectAction()  # ф-ия связи с эл-тами окна

    def _connectAction(self):
        self.ui.btnSaveTable.clicked.connect(lambda: self.saveTable())          #
        self.ui.btnAddStrInTable.clicked.connect(lambda: self.AddStrInTable())          #
        self.ui.btnDelStrLast.clicked.connect(lambda: self.delStrLast())           #
        self.ui.btnExitAndClose.clicked.connect(lambda: self.close())  #
        self.ui.btnSetNumPeopleInSquad.clicked.connect(lambda: self.setNumPeopleInSquad())  #

    def setNumPeopleInSquad(self):
        winNumSquads = QDialog()
        winNumSquads.ui = Ui_SetNumSquad() # вы
        winNumSquads.ui.setupUi(winNumSquads)
        winNumSquadsOut = winNumSquads.exec()
        
        if winNumSquadsOut == 1:  # если нажали да
            NumSquads = winNumSquads.ui.lineEditSetNumSquad.text() 
        else:  # иначе игнорируем
            return
        winTableNumPeopleInSquad = creatTableNumPeopleInSquad(self, int(NumSquads))
        winTableNumPeopleInSquad.exec_()
        

    def delStrLast(self):
        rowInTblTsk = self.ui.tableTaskVar.rowCount()
        sheet = self.book.active
        if rowInTblTsk > 0:
            self.ui.tableTaskVar.removeRow(rowInTblTsk-1)
        if sheet.max_row > 0:
            sheet.delete_cols(sheet.max_row,1)

    # def closeWinCreatTable(self):
    #     self.saveTable()
    #     self.close()

    def saveTable(self):
        sheet = self.book.active

        for rowInTblTsk in range(sheet.max_row):
            for colInTblTsk in range(sheet.max_column):
                sheet.cell(rowInTblTsk + 1, colInTblTsk + 1).value = None
        #row = []
        for rowInTblTsk in range(self.ui.tableTaskVar.rowCount()):
            for colInTblTsk in range(self.ui.tableTaskVar.columnCount()):
                if self.ui.tableTaskVar.item(rowInTblTsk, colInTblTsk):
                    tmpItem = self.ui.tableTaskVar.item(rowInTblTsk, colInTblTsk).text()
                else:
                    tmpItem = ' '
                sheet.cell(rowInTblTsk + 1, colInTblTsk + 1).value = tmpItem

        for rowInSqdTbl in range(len(self.listNumPeopleInSquad)):
            for colInSqdTbl in range(len(self.listNumPeopleInSquad[rowInSqdTbl])):
                sheet.cell(rowInSqdTbl + 1, colInSqdTbl + self.ui.tableTaskVar.columnCount()+1).value = self.listNumPeopleInSquad[rowInSqdTbl][colInSqdTbl]
                
        self.book.save(self.pathToExcelFile)

    def AddStrInTable(self): # генерируем строку в таблице для записи в нее чиселок
        #rowPosition = self.ui.tableTaskVar.rowCount()
        self.ui.tableTaskVar.insertRow(self.ui.tableTaskVar.rowCount() )  # вставляем в таблицу "строку таблицы из файла"
        #for colInTblTsk in range(self.ui.tableTaskVar.columnCount() + 1):
        #    self.ui.tableTaskVar.setItem(self.ui.tableTaskVar.rowCount() - 1, colInTblTsk, QtWidgets.QTableWidgetItem(' '))  # заполняем "строку таблицы из файла", каждую ячейку

    def openFile(self, pathToExcelFile): # открываем указанный файл в окне для редактирования вариантов
        self.pathToExcelFile = pathToExcelFile # сохраняем путь до файла
        # файлик с таблицой должен называться "В" + номер студента по списку + ".xlsx" (расширение файла)
        self.book = openpyxl.load_workbook(self.pathToExcelFile)  # открываем файл с помощью либы для обработки .xlsx
        sheet = self.book.active  # active - выбирает номер страницы в книге без параметров (по умолчанию) первая страница

        countColumns = 0 # счетчик колонок
        tabelVar = [] # список строк

        for row in sheet.iter_rows(sheet.min_row, sheet.max_row):  # подкачиваем данные из xlsx файла
            rowVar = []
            for cell in row: # Две последнии колонки обрезаются т к их некуда вписать !!!!!!!!!
                rowVar.append(cell.value)
            tabelVar.append(rowVar)

        for row in tabelVar:
            for cell in row:
                print(cell, "\t", sep="")
            print()

        self.ui.tableTaskVar.setRowCount(0)  # удаление старых данных из таблицы (если уже генерировалась таблица с заданием)

        for list in tabelVar:
            rowPosition = self.ui.tableTaskVar.rowCount()  # генерируем строку в таблице для записи в нее чиселок
            self.ui.tableTaskVar.insertRow(rowPosition)  # вставляем в таблицу "строку таблицы из файла"
            for item in list:
                if countColumns >= 0:
                    self.ui.tableTaskVar.setItem(rowPosition, countColumns, QtWidgets.QTableWidgetItem(item))  # заполняем "строку таблицы из файла", каждую ячейку
                countColumns = countColumns + 1
            countColumns = 0

    def closeEvent(self, event):
        close = QMessageBox()
        close.setWindowTitle("Закрыть редактор")
        close.setText("Вы уверены, что хотите закрыть редактор?")  #
        close.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)  #
        self.saveTable()
        close = close.exec()
        if close == QMessageBox.Ok:  # если нажали да
            self.book.close()
            event.accept()  # подтверждаем ивент
            #self.winEditTable.mainMenu.show()
        else:  # иначе игнорируем
            event.ignore()

class creatTableNumPeopleInSquad(QtWidgets.QDialog): # окно с таблицей количества людей в отделении
    def __init__(self, root, numPeopleInSquad):  # 
        """Initializer."""
        super().__init__(root)  # инициализация

        self.ui = Ui_winTableNumPeopleInSquad()  # инициализация ui
        self.ui.setupUi(self)  # инициализация ui окна (присвоение конкретных пар-ов)
        self.winEditTable = root  # сохраняем нашего родителя
        #self.listNumPeopleInSquad = []

        sizeWindow = QRect(QApplication.desktop().screenGeometry())  # смотрим размер экраны
        width = int(210*2+60)
        height = int(30*numPeopleInSquad + 120)
        #width = int(sizeWindow.width() - 2*(sizeWindow.width()) / 3)  # выставляем ширину окна
        #height = int(sizeWindow.height() - 2*(sizeWindow.height()) / 3)  # выставляем длину окна
        # присваиваем параметры длины и ширины окну
        self.resize(width, height)

        self.move(int(sizeWindow.width() / 20), int(sizeWindow.height() / 20))  # двигаем окно левее и выше

        quit = QAction("Quit", self)  # событие выхода
        quit.triggered.connect(self.closeEvent)  # если событие выхода срабатывает то вызывается closeEvent

        self.creatTable(numPeopleInSquad)

        self.openFile(self.winEditTable.pathToExcelFile)

        self._connectAction()  # ф-ия связи с эл-тами окна

    def _connectAction(self):
        self.ui.btnExitAndClose.clicked.connect(lambda: self.closeWinCreatTable())  #

    def closeWinCreatTable(self):
        self.saveTable()
        self.close()

    def openFile(self, pathToExcelFile): # открываем указанный файл в окне для редактирования вариантов
        self.pathToExcelFile = pathToExcelFile # сохраняем путь до файла
        # файлик с таблицой должен называться "В" + номер студента по списку + ".xlsx" (расширение файла)
        self.book = openpyxl.load_workbook(self.pathToExcelFile)  # открываем файл с помощью либы для обработки .xlsx
        sheet = self.book.active  # active - выбирает номер страницы в книге без параметров (по умолчанию) первая страница

        countColumns = 0 # счетчик колонок
        tabelVar = [] # список строк

        for row in sheet.iter_rows(sheet.min_row, sheet.max_row):  # подкачиваем данные из xlsx файла
            rowVar = []
            for cell in row: # Две последнии колонки обрезаются т к их некуда вписать !!!!!!!!!
                rowVar.append(cell.value)
            tabelVar.append(rowVar)

        # for row in tabelVar:
        #     for cell in row:
        #         print(cell, "\t", sep="")
        #     print()

        self.ui.tableTaskVar.setRowCount(0)  # удаление старых данных из таблицы (если уже генерировалась таблица с заданием)

        for list in tabelVar:
            rowPosition = self.ui.tableTaskVar.rowCount()  # генерируем строку в таблице для записи в нее чиселок
            self.ui.tableTaskVar.insertRow(rowPosition)  # вставляем в таблицу "строку таблицы из файла"
            for item in list:
                if countColumns >= 0:
                    self.ui.tableTaskVar.setItem(rowPosition, countColumns, QtWidgets.QTableWidgetItem(item))  # заполняем "строку таблицы из файла", каждую ячейку
                countColumns = countColumns + 1
            countColumns = 0

    def saveTable(self):
        self.winEditTable.listNumPeopleInSquad = []
        for rowInTblTsk in range(self.ui.tableNumPeopleInSquad.rowCount()):
            self.winEditTable.listNumPeopleInSquad.append([])
            for colInTblTsk in range(self.ui.tableNumPeopleInSquad.columnCount()):
                if self.ui.tableNumPeopleInSquad.item(rowInTblTsk, colInTblTsk):
                    tmpItem = int(self.ui.tableNumPeopleInSquad.item(rowInTblTsk, colInTblTsk).text())
                    
                else:
                    tmpItem = 0
                #print("item", tmpItem)
                self.winEditTable.listNumPeopleInSquad[-1].append(tmpItem)
        #self.winEditTable.listNumPeopleInSquad = self.listNumPeopleInSquad


    def closeEvent(self, event):
        close = QMessageBox()
        close.setWindowTitle("Закрыть окно")
        close.setText("Вы уверены, что хотите закрыть окно?")  #
        close.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)  #
        close = close.exec()
        if close == QMessageBox.Ok:  # если нажали да
            # self.book.close()
            event.accept()  # подтверждаем ивент
            self.winEditTable.show()
            #self.winEditTable.mainMenu.show()
        else:  # иначе игнорируем
            event.ignore()

    def creatTable(self, numPeopleInSquad):
        self.ui.tableNumPeopleInSquad.setRowCount(0)  # удаление старых данных из таблицы (если уже генерировалась таблица с заданием)

        for i in range(numPeopleInSquad):
            rowPosition = self.ui.tableNumPeopleInSquad.rowCount()  # генерируем строку в таблице для записи в нее чиселок
            self.ui.tableNumPeopleInSquad.insertRow(rowPosition)  # вставляем в таблицу "строку таблицы из файла"
            self.ui.tableNumPeopleInSquad.setItem(rowPosition, 0, QtWidgets.QTableWidgetItem(str(i + 1)))  # заполняем "строку таблицы из файла", каждую ячейку
            

class winSearchKey(QtWidgets.QDialog): # окно для загрузки ключа преподавателя при первом запуске программы

    def __init__(self, root): # передаем параметр root это родитель т е MainMenu (в этом классе и лежит наше окно winSigReport)
        """Initializer."""
        super().__init__(root) # инициализация

        self.ui = Ui_first_launch() # инициализация ui
        self.ui.setupUi(self) # инициализация ui окна (присвоение конкретных пар-ов)
        self.mainMenu = root  # сохраняем нашего родителя


        sizeWindow = QRect(QApplication.desktop().screenGeometry())         # смотрим размер экраны
        width = int(sizeWindow.width() - (sizeWindow.width()) * 2 / 3)      # выставляем ширину окна
        height = int(sizeWindow.height() - (sizeWindow.height()) * 2 / 3)   # выставляем длину окна
        # присваиваем параметры длины и ширины окну
        self.resize(width, height)

        self.move(int(sizeWindow.width() / 20), int(sizeWindow.height() / 20)) # двигаем окно левее и выше

        basedir = os.path.dirname(__file__)
        self.encrypted_data_path = self.join(basedir, "encrypted_data")
        self.first_launch_txt_path = self.join(basedir, "first_launch", "first_launch.txt")

        quit = QAction("Quit", self)  # событие выхода
        quit.triggered.connect(self.closeEvent)  # если событие выхода срабатывает то вызывается closeEvent

        self.ui.btnSearchPathToKey.setCheckable(True)

        self._connectAction() # ф-ия связи с эл-тами окна

    def _connectAction(self):
        self.ui.btnSearchPathToKey.clicked.connect(lambda: self.select_key()) # прописываем действие по кнопке

    def select_key(self):
        file_name = QtWidgets.QFileDialog.getOpenFileName()[0]
        if file_name == "":
            return

        if not os.path.exists(file_name):
            message_box_create("Первоначальная дешифровка файлов", "Выбранный ключ-файл не существует",
                               QMessageBox.Critical)
            return

        try:
            with open(file_name, "rb") as file:
                key = file.read()
        except Exception:
            message_box_create("Первоначальная дешифровка файлов", "Выбранный ключ-файл повреждён",
                               QMessageBox.Critical)
            return

        try:
            found_files = find_files(Path(self.encrypted_data_path))
        except Exception:
            message_box_create("Первоначальная дешифровка файлов", "Не удалось дешифровать файлы программы",
                               QMessageBox.Critical)
            return

        for file in found_files:
            try:
                content = initial_decrypt_file(file, key.decode())
            except Exception:
                message_box_create("Первоначальная дешифровка файлов",
                                   "Выбранный ключ-файл не подходит для дешифровки файлов", QMessageBox.Critical)
                return
            if content == b"ERROR_DECRYPT":
                message_box_create("Первоначальная дешифровка файлов",
                                   "Выбранный ключ-файл не подходит для дешифровки файлов", QMessageBox.Critical)
                return
            nonce, cipher_content, tag = aes_encrypt(content, aes_generate_key())
            try:
                with open(file.resolve(), "wb") as output_file:
                    output_file.write(nonce)
                    output_file.write(tag)
                    output_file.write(cipher_content)
            except Exception:
                message_box_create("Первоначальная дешифровка файлов", "Не удалось зашифровать файлы программы",
                                   QMessageBox.Critical)
                return
        try:
            with open(self.first_launch_txt_path, "w") as fd:
                fd.write("false")
        except Exception:
            message_box_create("Первоначальная дешифровка файлов",
                               "Файлы программы повреждены. Необходимо переустановить программу", QMessageBox.Critical)
        self.close()

    def join(self,*args):
        return os.path.join(*args).replace(os.path.sep, "/")
    
    def closeEvent(self, event):
        if self.ui.btnSearchPathToKey.isChecked():
            event.accept()  # подтверждаем ивент
            return
        close = QMessageBox()
        close.setWindowTitle("Закрыть окно")
        close.setText("Вы уверены, что хотите закрыть окно?")  #
        close.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)  #
        close = close.exec()
        if close == QMessageBox.Ok:  # если нажали да
            event.accept()  # подтверждаем ивент
            if not( self.ui.btnSearchPathToKey.isChecked()):
                sys.exit()
        else:  # иначе игнорируем
            event.ignore()
        self.ui.btnSearchPathToKey.setChecked(False)



